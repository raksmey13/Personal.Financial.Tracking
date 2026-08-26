import csv
import io
import os
from fastapi import APIRouter, HTTPException, UploadFile, File, Query, status, Depends
from fastapi.responses import StreamingResponse
from typing import Optional
from datetime import datetime, date
from decimal import Decimal
from sqlmodel import select, func, extract
from database import SessionDep
from models import Transaction, Account, Category, User

# --- EXCEL GENERATION IMPORTS ---
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# --- REPORTLAB IMPORTS FOR PDF GENERATION ---
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

from .auth import get_current_user

router = APIRouter(prefix="/export-import", tags=["Export & Import"])

# Register Khmer Font for PDF generation if available, fallback gracefully
KHMER_FONT_NAME = "Helvetica"
possible_khmer_fonts = [
    "fonts/NotoSansKhmer-Regular.ttf",
    "backend/fonts/NotoSansKhmer-Regular.ttf",
    "/usr/share/fonts/truetype/noto/NotoSansKhmer-Regular.ttf"
]
for font_path in possible_khmer_fonts:
    if os.path.exists(font_path):
        try:
            pdfmetrics.registerFont(TTFont('NotoSansKhmer', font_path))
            KHMER_FONT_NAME = 'NotoSansKhmer'
            break
        except Exception:
            pass


# =========================================================
# 1A. DOWNLOAD EXCEL IMPORT TEMPLATE (.XLSX)
# =========================================================
@router.get("/template/excel")
def download_excel_template():
    """Generates a formatted multi-sheet Excel template for bulk transaction imports."""
    try:
        wb = Workbook()

        # Styles Setup
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="3D5AFE", end_color="3D5AFE", fill_type="solid")
        info_header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")

        thin_border = Border(
            left=Side(style="thin", color="E2E8F0"),
            right=Side(style="thin", color="E2E8F0"),
            top=Side(style="thin", color="E2E8F0"),
            bottom=Side(style="thin", color="E2E8F0")
        )
        align_center = Alignment(horizontal="center", vertical="center")

        # SHEET 1: Data Entry Template (First Tab for Easy Access)
        ws_tx = wb.active
        ws_tx.title = "Transactions"
        ws_tx.views.sheetView[0].showGridLines = True

        headers = ["date", "account_name", "category_name", "type", "amount", "description"]
        ws_tx.append(headers)

        for cell in ws_tx[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center

        sample_rows = [
            ["2026-08-15", "ABA USD", "Food & Dining", "expense", 12.50, "Lunch meeting"],
            ["2026-08-15", "ABA KHR", "Transport", "expense", 6000, "Tuktuk ride"],
            ["2026-08-01", "ABA USD", "Salary", "income", 1200.00, "Monthly Paycheck"]
        ]

        for row in sample_rows:
            ws_tx.append(row)

        # Dropdown Validation for 'type' Column
        dv_type = DataValidation(type="list", formula1='"income,expense,transfer"', allow_blank=False)
        ws_tx.add_data_validation(dv_type)
        dv_type.add("D2:D1000")

        # SHEET 2: Instructions & Field Rules
        ws_info = wb.create_sheet(title="Instructions")
        ws_info.views.sheetView[0].showGridLines = True

        ws_info.append(["Field Name", "Required?", "Allowed Format / Values", "Description & Examples"])
        instructions_data = [
            ["date", "Yes", "YYYY-MM-DD (e.g. 2026-08-15)", "The date when the transaction occurred."],
            ["account_name", "Yes", "Text (e.g. ABA USD, Cash)", "Must match an existing active account in your ledger."],
            ["category_name", "Yes", "Text (e.g. Food & Dining, Salary)", "Category or subcategory title."],
            ["type", "Yes", "income | expense | transfer", "Classification of cash flow."],
            ["amount", "Yes", "Numeric (e.g. 15.50 or 6000)", "Amount in native account currency."],
            ["description", "No", "Text", "Optional notes or merchant description."]
        ]

        for row in instructions_data:
            ws_info.append(row)

        for cell in ws_info[1]:
            cell.font = header_font
            cell.fill = info_header_fill
            cell.alignment = align_center

        # Apply Width Auto-fitting and Borders
        for sheet in [ws_tx, ws_info]:
            for col in sheet.columns:
                max_len = max(len(str(cell.value or "")) for cell in col)
                col_letter = get_column_letter(col[0].column)
                sheet.column_dimensions[col_letter].width = max(max_len + 4, 15)

                for cell in col:
                    if cell.row > 1:
                        cell.border = thin_border

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=netstream_import_template.xlsx"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to generate Excel template: {str(e)}")


# =========================================================
# 1B. DOWNLOAD CSV IMPORT TEMPLATE
# =========================================================
@router.get("/template")
def download_template():
    """Generates a structured CSV template matching model.py specs for bulk importing."""
    output = io.StringIO()
    writer = csv.writer(output)

    writer.writerow(["date", "account_name", "category_name", "type", "amount", "description"])
    writer.writerow(["2026-08-01", "ABA USD", "Food & Dining", "expense", "15.50", "Lunch meeting"])
    writer.writerow(["2026-08-01", "ABA USD", "Salary", "income", "1200.00", "Monthly Paycheck"])
    writer.writerow(["2026-08-01", "ABA USD", "Sweep Saving", "transfer", "100.00", "50/30/20 Savings Allocation"])

    output.seek(0)
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode("utf-8")),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=netstream_import_template.csv"}
    )


# =========================================================
# 2. EXPORT TRANSACTIONS TO CSV
# =========================================================
@router.get("/csv")
def export_csv(
        session: SessionDep,
        current_user: User = Depends(get_current_user),
        start_date: Optional[date] = Query(None),
        end_date: Optional[date] = Query(None),
        account_id: Optional[int] = Query(None),
        category_type: Optional[str] = Query("both"),
        include_income: bool = Query(True),
        include_expense: bool = Query(True),
        include_transfer: bool = Query(True)
):
    """Queries PostgreSQL using current user context and streams a filtered CSV file."""
    statement = select(Transaction).where(Transaction.user_id == current_user.id)

    if start_date:
        statement = statement.where(Transaction.transaction_date >= start_date)
    if end_date:
        statement = statement.where(Transaction.transaction_date <= end_date)

    if account_id:
        statement = statement.where(Transaction.account_id == account_id)

    allowed_types = []
    if category_type == "income" or (category_type == "both" and include_income):
        allowed_types.append("income")
    if category_type == "expense" or (category_type == "both" and include_expense):
        allowed_types.append("expense")
    if category_type == "both" and include_transfer:
        allowed_types.append("transfer")

    if allowed_types:
        statement = statement.where(func.lower(Transaction.type).in_(allowed_types))

    statement = statement.order_by(Transaction.transaction_date.desc(), Transaction.id.desc())
    transactions = session.exec(statement).all()

    output = io.StringIO()
    writer = csv.writer(output)

    writer.writerow(["ID", "Date", "Account Name", "Category Name", "Type", "Amount", "Description"])

    for tx in transactions:
        account = session.get(Account, tx.account_id)
        category = session.get(Category, tx.category_id) if tx.category_id else None

        acc_name = account.account_name if account else "Unknown Account"
        cat_name = category.name if category else "Uncategorized"

        raw_amount = abs(tx.amount)
        formatted_amount = f"-{raw_amount}" if tx.type.lower() == "expense" else f"{raw_amount}"

        writer.writerow([
            tx.id,
            tx.transaction_date,
            acc_name,
            cat_name,
            tx.type,
            formatted_amount,
            tx.description or ""
        ])

    output.seek(0)
    filename = f"ledger_export_{date.today().strftime('%Y%m%d')}.csv"

    return StreamingResponse(
        io.BytesIO(output.getvalue().encode("utf-8")),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# =========================================================
# 3. EXPORT TRANSACTIONS TO PDF
# =========================================================
@router.get("/pdf")
def export_pdf(
        session: SessionDep,
        current_user: User = Depends(get_current_user),
        period: Optional[str] = Query("2026"),
        account_id: Optional[int] = Query(None),
        include_income: bool = Query(True),
        include_expense: bool = Query(True),
        include_transfer: bool = Query(True),
        two_column: bool = Query(False)
):
    """Generates and streams a formatted PDF report matching current_user context."""
    statement = select(Transaction).where(Transaction.user_id == current_user.id)

    if period and period.isdigit():
        statement = statement.where(extract('year', Transaction.transaction_date) == int(period))

    if account_id:
        statement = statement.where(Transaction.account_id == account_id)

    allowed_types = []
    if include_income:
        allowed_types.append("income")
    if include_expense:
        allowed_types.append("expense")
    if include_transfer:
        allowed_types.append("transfer")

    if allowed_types:
        statement = statement.where(func.lower(Transaction.type).in_(allowed_types))

    statement = statement.order_by(Transaction.transaction_date.desc(), Transaction.id.desc())
    transactions = session.exec(statement).all()

    pdf_buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        pdf_buffer,
        pagesize=letter,
        rightMargin=36,
        leftMargin=36,
        topMargin=36,
        bottomMargin=36
    )

    styles = getSampleStyleSheet()
    story = []

    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontSize=18,
        leading=22,
        fontName='Helvetica-Bold',
        textColor=colors.HexColor('#1E293B'),
        spaceAfter=12
    )
    cell_style = ParagraphStyle(
        'CellText',
        parent=styles['Normal'],
        fontSize=8,
        leading=10,
        fontName=KHMER_FONT_NAME if KHMER_FONT_NAME != 'Helvetica' else 'Helvetica',
        textColor=colors.HexColor('#1E293B')
    )
    cell_header_style = ParagraphStyle(
        'HeaderCellText',
        parent=styles['Normal'],
        fontSize=9,
        leading=11,
        fontName='Helvetica-Bold',
        textColor=colors.whitesmoke
    )

    story.append(Paragraph(f"Transaction Summary Report ({period})", title_style))
    story.append(Spacer(1, 10))

    headers = ["Date", "Account", "Category", "Type", "Amount", "Description"]
    table_data = [[Paragraph(h, cell_header_style) for h in headers]]

    for tx in transactions:
        account = session.get(Account, tx.account_id)
        category = session.get(Category, tx.category_id) if tx.category_id else None

        acc_name = account.account_name if account else "N/A"
        curr = str(account.currency if account else "USD").strip().upper()
        cat_name = category.name if category else "Uncategorized"

        # Sanitize Khmer text & currency symbol to prevent PDF font crashes ("tofu" blocks)
        raw_desc = tx.description or ""
        if KHMER_FONT_NAME == "Helvetica":
            # Strip non-latin characters and replace ៛ with KHR if custom font isn't loaded
            raw_desc = "".join([c for c in raw_desc if ord(c) < 128])
            symbol = " KHR" if curr == "KHR" else "$"
        else:
            symbol = "៛" if curr == "KHR" else "$"

        raw_amount = abs(tx.amount)
        if curr == "KHR":
            symbol_fmt = f"{raw_amount:,.0f}{symbol}"
        else:
            symbol_fmt = f"{symbol}{raw_amount:,.2f}"

        formatted_amount = f"-{symbol_fmt}" if tx.type.lower() == "expense" else f"+{symbol_fmt}"

        table_data.append([
            Paragraph(str(tx.transaction_date), cell_style),
            Paragraph(acc_name, cell_style),
            Paragraph(cat_name, cell_style),
            Paragraph(tx.type.capitalize(), cell_style),
            Paragraph(formatted_amount, cell_style),
            Paragraph(raw_desc, cell_style)
        ])

    col_widths = [65, 85, 95, 55, 75, 165]

    pdf_table = Table(table_data, colWidths=col_widths)
    pdf_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#5C6BC0')),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
    ]))

    story.append(pdf_table)
    doc.build(story)

    pdf_buffer.seek(0)
    filename = f"transaction_summary_{period}.pdf"

    return StreamingResponse(
        pdf_buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# =========================================================
# 4. UNIVERSAL IMPORT (CSV & EXCEL .XLSX SUPPORTED)
# =========================================================
@router.post("/import")
async def import_file(
    session: SessionDep,
    current_user: User = Depends(get_current_user),
    file: UploadFile = File(...)
):
    """
    Parses uploaded CSV or Excel (.xlsx) files under current_user context.
    """
    filename_lower = file.filename.lower()
    is_excel = filename_lower.endswith('.xlsx') or filename_lower.endswith('.xls')
    is_csv = filename_lower.endswith('.csv') or filename_lower.endswith('.txt')

    if not (is_excel or is_csv):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid file format. Please upload a .csv or .xlsx file."
        )

    contents = await file.read()
    parsed_rows = []

    if is_excel:
        try:
            wb = load_workbook(filename=io.BytesIO(contents), data_only=True)
            ws = wb["Transactions"] if "Transactions" in wb.sheetnames else wb.active

            headers = [str(cell.value or "").strip().lower() for cell in ws[1]]
            for row_cells in ws.iter_rows(min_row=2, values_only=True):
                if not any(row_cells):
                    continue
                row_dict = dict(zip(headers, row_cells))
                parsed_rows.append(row_dict)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Failed to read Excel file: {str(e)}")

    else:
        try:
            decoded = contents.decode("utf-8")
        except UnicodeDecodeError:
            decoded = contents.decode("latin-1")

        csv_reader = csv.DictReader(io.StringIO(decoded))
        for row in csv_reader:
            parsed_rows.append({k.strip().lower(): v for k, v in row.items() if k})

    imported_count = 0
    errors = []

    fallback_category = session.exec(
        select(Category).where(Category.user_id == current_user.id)
    ).first()

    for line_num, row in enumerate(parsed_rows, start=2):
        try:
            raw_date = str(row.get("date", "")).strip()
            raw_account = str(row.get("account_name", "")).strip()
            raw_category = str(row.get("category_name", "")).strip()
            raw_type = (str(row.get("type", "")).strip() or "expense").lower()
            raw_amount = str(row.get("amount", "0")).strip().replace("$", "").replace(",", "").replace("៛", "")
            raw_desc = str(row.get("description", "")).strip()

            if raw_desc.lower() == "none":
                raw_desc = ""

            amount_val = Decimal(raw_amount)
            if amount_val == 0:
                continue

            tx_date = date.today()
            if raw_date and raw_date.lower() != "none":
                if isinstance(row.get("date"), (datetime, date)):
                    tx_date = row.get("date").date() if isinstance(row.get("date"), datetime) else row.get("date")
                else:
                    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d"):
                        try:
                            tx_date = datetime.strptime(raw_date.split()[0], fmt).date()
                            break
                        except ValueError:
                            pass

            account = session.exec(
                select(Account).where(
                    func.lower(Account.account_name) == func.lower(raw_account),
                    Account.user_id == current_user.id,
                    Account.is_active == True
                )
            ).first()

            if not account:
                errors.append(f"Row {line_num}: Account '{raw_account}' does not exist in your active accounts.")
                continue

            category_id = fallback_category.id if fallback_category else None
            if raw_category and raw_category.lower() != "none":
                matched_category = session.exec(
                    select(Category).where(
                        func.lower(Category.name) == func.lower(raw_category)
                    )
                ).first()
                if matched_category:
                    category_id = matched_category.id

            clean_amount = abs(amount_val)

            new_tx = Transaction(
                user_id=current_user.id,
                account_id=account.id,
                category_id=category_id,
                transaction_date=tx_date,
                type=raw_type,
                amount=clean_amount,
                description=raw_desc or "Bulk File Import"
            )

            if raw_type == "income":
                account.balance += clean_amount
            elif raw_type in ["expense", "transfer"]:
                account.balance -= clean_amount

            session.add(new_tx)
            session.add(account)
            imported_count += 1

        except Exception as err:
            errors.append(f"Row {line_num}: Failed to parse -> {str(err)}")

    if imported_count > 0:
        session.commit()

    return {
        "status": 200,
        "message": f"Successfully imported {imported_count} transactions.",
        "errors": errors
    }